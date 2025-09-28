"""
Excel Exporter Module
Handles exporting data in various formats with formatting.
"""

import streamlit as st
import pandas as pd
from typing import Dict, Any, Optional
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExporter:
    """Handles data export operations with formatting."""
    """Handles data export operations with formatting."""
    
    def __init__(self):
        """Initialize the exporter with default styles."""
        self.default_styles = {
            'header': {
                'fill': PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid'),
                'font': Font(bold=True),
                'alignment': Alignment(vertical='top', wrap_text=True),
                'border': Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            },
            'cell': {
                'alignment': Alignment(vertical='top', wrap_text=True)
            },
            'error': {
                'fill': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                'font': Font(color='9C0006')
            },
            'warning': {
                'fill': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                'font': Font(color='9C6500')
            }
        }
    
    def show_export_interface(
        self,
        df: pd.DataFrame,
        validation_results: Optional[Dict[str, Any]] = None
    ) -> None:
        """
        Display the export interface in Streamlit.
        
        Args:
            df: Dataframe to export
            validation_results: Optional validation results to highlight issues
        """
        st.write("### Export Data")
        
        # Export options
        st.write("#### Export Options")
        
        # Export format selection
        export_format = st.selectbox(
            "Select Export Format",
            options=["Excel (XLSX)", "CSV", "Excel (XLS)", "TSV", "JSON", "XML"],
            help="Choose the format for your exported data"
        )
        
        # File name with dynamic extension
        default_extension = {
            "Excel (XLSX)": ".xlsx",
            "CSV": ".csv",
            "Excel (XLS)": ".xls",
            "TSV": ".tsv",
            "JSON": ".json",
            "XML": ".xml"
        }[export_format]
        
        filename = st.text_input(
            "Output filename",
            value=f"standardized_data{default_extension}",
            help=f"Enter the name for your output file (will be saved as {default_extension})"
        )
        
        # Sheet name (only for Excel formats)
        sheet_name = "Standardized Data"  # Default sheet name
        if "Excel" in export_format:
            sheet_name = st.text_input(
                "Sheet name",
                value=sheet_name,
                help="Enter the name for the worksheet"
            )
        
        # Formatting options
        st.write("#### Formatting Options")
        col1, col2 = st.columns(2)
        
        with col1:
            include_validation = st.checkbox(
                "Include validation results",
                value=True,
                help="Highlight cells with validation issues"
            )
        
        with col2:
            freeze_header = st.checkbox(
                "Freeze header row",
                value=True,
                help="Keep header row visible while scrolling"
            )
        
        # Export button
        if st.button("Export Data"):
            # Get the appropriate mime type
            mime_types = {
                "Excel (XLSX)": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "Excel (XLS)": "application/vnd.ms-excel",
                "CSV": "text/csv",
                "TSV": "text/tab-separated-values",
                "JSON": "application/json",
                "XML": "application/xml"
            }
            
            # Default sheet name for Excel formats
            sheet_name_val = sheet_name if "Excel" in export_format else "Sheet1"
            
            # Create the export file
            file_data = self._create_export_file(
                df,
                validation_results if include_validation else None,
                sheet_name_val,
                freeze_header,
                export_format
            )
            
            # Show download button
            st.download_button(
                label=f"📥 Download {export_format} file",
                data=file_data,
                file_name=filename,
                mime=mime_types[export_format]
            )
    
    def _create_export_file(
        self,
        df: pd.DataFrame,
        validation_results: Optional[Dict[str, Any]],
        sheet_name: str,
        freeze_header: bool,
        export_format: str
    ) -> bytes:
        """
        Create formatted export file.
        
        Args:
            df: Dataframe to export
            validation_results: Optional validation results
            sheet_name: Name of the worksheet (for Excel formats)
            freeze_header: Whether to freeze the header row (for Excel formats)
            export_format: The chosen export format
            
        Returns:
            File contents as bytes
        """
        # Create output buffer
        output = io.BytesIO()
        
        # Clean up the dataframe before export
        export_df = df.copy()
        
        # Handle data types carefully to prevent data loss
        for col in export_df.columns:
            # Get a sample of non-null values to check actual content
            sample = export_df[col].dropna().head()
            
            if export_df[col].dtype == 'object':
                # Check if column contains numbers stored as strings
                try:
                    if all(str(x).replace('.','',1).isdigit() for x in sample if pd.notna(x)):
                        # If they're all numbers, preserve them as numeric
                        export_df[col] = pd.to_numeric(export_df[col], errors='coerce')
                    else:
                        # If mixed content or pure strings, preserve as strings
                        export_df[col] = export_df[col].fillna('').astype(str)
                except:
                    # If any error, safely convert to string
                    export_df[col] = export_df[col].fillna('').astype(str)
            elif 'int' in str(export_df[col].dtype):
                # Preserve integers exactly as they are
                pass
            elif 'float' in str(export_df[col].dtype):
                # Preserve floats exactly as they are
                pass
                
        try:
            if export_format == "CSV":
                export_df.to_csv(output, index=False)
            elif export_format == "TSV":
                export_df.to_csv(output, index=False, sep='\t')
            elif export_format == "JSON":
                export_df.to_json(output, orient='records', indent=2)
            elif export_format == "XML":
                # Convert DataFrame to XML
                xml_content = self._dataframe_to_xml(export_df)
                output.write(xml_content.encode('utf-8'))
            else:  # Excel formats
                if "Excel (XLSX)" in export_format:
                    try:
                        with pd.ExcelWriter(
                            output,
                            engine='openpyxl',
                            mode='w'
                        ) as writer:
                            # Write the data
                            export_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            # Apply basic formatting with openpyxl
                            try:
                                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                                from openpyxl.utils import get_column_letter
                                
                                worksheet = writer.sheets[sheet_name]
                                
                                # Apply header styles
                                header_style = self.default_styles['header']
                                for col in range(1, len(export_df.columns) + 1):
                                    cell = worksheet.cell(row=1, column=col)
                                    cell.fill = header_style['fill']
                                    cell.font = header_style['font']
                                    cell.alignment = header_style['alignment']
                                    cell.border = header_style['border']
                                    
                                # Apply cell styles to data area
                                cell_style = self.default_styles['cell']
                                for row in range(2, len(export_df) + 2):
                                    for col in range(1, len(export_df.columns) + 1):
                                        cell = worksheet.cell(row=row, column=col)
                                        cell.alignment = cell_style['alignment']
                                        
                                # Apply validation highlighting if provided
                                if validation_results:
                                    error_style = self.default_styles['error']
                                    warning_style = self.default_styles['warning']
                                    
                                    for col, issues in validation_results.get('column_issues', {}).items():
                                        if col in export_df.columns:
                                            col_idx = list(export_df.columns).index(col) + 1
                                            for row_idx, issue in enumerate(issues, start=2):
                                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                                if 'error' in issue.get('level', '').lower():
                                                    cell.fill = error_style['fill']
                                                    cell.font = error_style['font']
                                                elif 'warning' in issue.get('level', '').lower():
                                                    cell.fill = warning_style['fill']
                                                    cell.font = warning_style['font']
                                
                                # Freeze header if requested
                                if freeze_header:
                                    worksheet.freeze_panes = 'A2'
                                
                                # Auto-adjust column widths
                                for col in range(1, len(export_df.columns) + 1):
                                    letter = get_column_letter(col)
                                    max_length = 0
                                    column = [cell.value for cell in worksheet[letter]]
                                    
                                    for cell in column:
                                        try:
                                            max_length = max(max_length, len(str(cell)))
                                        except TypeError:
                                            pass
                                    
                                    adjusted_width = min(max_length + 2, 50)
                                    worksheet.column_dimensions[letter].width = adjusted_width
                                
                            except Exception as e:
                                st.warning(f"Could not apply Excel formatting: {str(e)}")
                    except Exception as e:
                        st.error(f"Error creating Excel file: {str(e)}")
                        return b''
            
            # Read back for verification
            output.seek(0)
            
            # Success message
            st.success("✅ File created successfully!")
            
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error creating file: {str(e)}")
            return b''
    
    def _dataframe_to_xml(self, df: pd.DataFrame) -> str:
        """Convert DataFrame to XML format."""
        xml_lines = ['<?xml version="1.0" encoding="UTF-8"?>']
        xml_lines.append('<data>')
        
        for index, row in df.iterrows():
            xml_lines.append(f'  <record id="{index}">')
            for col, value in row.items():
                # Escape XML special characters
                if pd.isna(value):
                    value = ""
                else:
                    value = str(value).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                xml_lines.append(f'    <{col}>{value}</{col}>')
            xml_lines.append('  </record>')
        
        xml_lines.append('</data>')
        return '\n'.join(xml_lines)
        
    
    def _apply_validation_highlighting(
        self,
        worksheet: Any,
        validation_results: Dict[str, Any],
        formats: Dict[str, Any],
        df: pd.DataFrame
    ) -> None:
        """
        Apply conditional formatting based on validation results.
        
        Args:
            worksheet: Excel worksheet object
            validation_results: Validation results dictionary
            formats: Dictionary of Excel formats
            df: Source dataframe
        """
        if not validation_results:
            return
            
        # Create error and warning cell formats
        error_format = formats['error']
        warning_format = formats['warning']
        
        # Apply error formatting
        for error in validation_results.get('errors', []):
            if not isinstance(error, dict) or 'rows' not in error:
                continue
                
            # If column is specified, apply to specific column only
            if 'column' in error:
                column = error['column']
                try:
                    col_idx = list(df.columns).index(column)
                    for row in error['rows']:
                        worksheet.write(row + 1, col_idx, df.iloc[row, col_idx], error_format)
                except (ValueError, IndexError):
                    continue
            # If no column specified, apply to entire row
            else:
                for row in error['rows']:
                    for col in range(len(df.columns)):
                        worksheet.write(row + 1, col, df.iloc[row, col], error_format)
        
        # Apply warning formatting
        for warning in validation_results.get('warnings', []):
            if not isinstance(warning, dict) or 'rows' not in warning:
                continue
                
            # If column is specified, apply to specific column only
            if 'column' in warning:
                column = warning['column']
                try:
                    col_idx = list(df.columns).index(column)
                    for row in warning['rows']:
                        worksheet.write(row + 1, col_idx, df.iloc[row, col_idx], warning_format)
                except (ValueError, IndexError):
                    continue
            # If no column specified, apply to entire row
            else:
                for row in warning['rows']:
                    for col in range(len(df.columns)):
                        worksheet.write(row + 1, col, df.iloc[row, col], warning_format)
